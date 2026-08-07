# -*- coding: utf-8 -*-
"""실파일 회귀 — 실제 최신 분석 xlsx가 로컬에 있을 때만 실행(없으면 SKIP).

검증:
  1. read_integrated → compute_pivots 내부 정합(sum(월합)==총계, 총계>0).
  2. 임시 복사본에 합성 6월을 병합해도 1~5월 합계가 불변('5월까지 불변').
     - 순수 파이썬 레벨(compute_pivots 월합) + 파생시트 수식 레벨(1~5월 컬럼 byte-identical).
  3. _apply_pivots가 SUMIFS 수식을 파괴하지 않고 신규 월 컬럼만 추가(멱등).
"""
import unittest, importlib.util, tempfile
from pathlib import Path
import openpyxl

FOLDER = Path('G:/공유 드라이브/인사총무팀_일반/05_총무 영역/비품·고정비')
SRC = Path(__file__).resolve().parents[1] / 'templates/scripts/slack-jipsa'


def load(n):
    spec = importlib.util.spec_from_file_location(n, SRC / f'{n}.py')
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


@unittest.skipUnless(FOLDER.exists(), '실 Drive 폴더 없음')
class Regression(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.ps = load('purchase_store')
        cls.pu = load('purchase')
        cls.analysis, _ = cls.ps.latest_analysis(FOLDER, '비품주문분석')
        assert cls.analysis is not None, '분석 파일을 찾지 못함'

    def _months(self):
        rows = self.ps.read_integrated(self.analysis)
        return rows, self.pu.compute_pivots(rows)

    def test_pivot_matches_dashboard_total(self):
        rows, p = self._months()
        self.assertEqual(sum(p['월합'].values()), p['총계'])   # 내부 정합
        self.assertGreater(p['총계'], 0)

    def test_no_charts_or_images_in_analysis_file(self):
        """분석 워크북은 차트/이미지가 없는 셀·수식 기반이어야 한다(openpyxl 병합이
        무손실이라는 전제의 트립와이어). 누군가 네이티브 차트/이미지를 추가하면
        write_merged_analysis의 openpyxl load→save 왕복에서 유실되므로 이 테스트가
        실패해 알려준다."""
        wb = openpyxl.load_workbook(self.analysis)
        n_charts = sum(len(ws._charts) for ws in wb.worksheets)
        n_images = sum(len(ws._images) for ws in wb.worksheets)
        self.assertEqual(n_charts, 0,
                          '분석 파일에 차트가 있음 — openpyxl 병합(write_merged_analysis) 시 유실 위험')
        self.assertEqual(n_images, 0,
                          '분석 파일에 이미지가 있음 — openpyxl 병합(write_merged_analysis) 시 유실 위험')

    def test_merge_preserves_months_through_may(self):
        rows, p0 = self._months()
        # 합성 신규 월(가장 큰 월 + 1) 한 건
        maxm = max(int(k[:-1]) for k in p0['월합'])
        newm = f'{maxm + 1}월'
        new = [{'월': newm, '일자': '3일', '부서': '인사총무', '용도': '사내비품',
                '카테고리': '사무용품', '품목': '_회귀합성_', '금액': 12345,
                '블록ID': f'{newm}#TEST'}]
        piv = self.pu.compute_pivots(rows + new)

        # (1) 순수 파이썬: 기존 월 월합 불변
        for m, v in p0['월합'].items():
            self.assertEqual(piv['월합'][m], v, f'{m} 월합이 변함')
        self.assertEqual(piv['월합'][newm], 12345)

        src_wb = openpyxl.load_workbook(self.analysis, data_only=False)
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / 'out.xlsx'
            self.ps.write_merged_analysis(self.analysis, out, new, pivots=piv)
            wf = openpyxl.load_workbook(out, data_only=False)

            # (2) 통합원본 append 확인(원본 값 보존 + 신규행 추가)
            si = openpyxl.load_workbook(self.analysis, data_only=True)[self.ps.INTEGRATED_SHEET]
            oi = openpyxl.load_workbook(out, data_only=True)[self.ps.INTEGRATED_SHEET]
            self.assertEqual(oi.max_row, si.max_row + len(new))
            self.assertEqual(oi.cell(oi.max_row, 1).value, newm)

            # (3) 파생 월 시트: 기존 월 컬럼 수식 byte-identical + 신규 월 헤더 추가
            for sheet, lay in self.ps.PIVOT_LAYOUT.items():
                if lay['axis'] != 'month':
                    continue
                s0, s1 = src_wb[sheet], wf[sheet]
                hr, fr, fmc = lay['header_row'], lay['first_row'], lay['first_month_col']
                # 기존 월 헤더 개수
                mcount = 0
                c = fmc
                while self.ps._month_num(s0.cell(hr, c).value) is not None:
                    mcount += 1
                    c += 1
                # 합계행(원본)
                r = fr
                while str(s0.cell(r, 1).value).strip() != '합계':
                    r += 1
                tot0 = r
                # 기존 월 컬럼(전 데이터행) 수식 동일
                for rr in range(fr, tot0):
                    for cc in range(fmc, fmc + mcount):
                        self.assertEqual(s0.cell(rr, cc).value, s1.cell(rr, cc).value,
                                         f'{sheet} r{rr}c{cc} 기존 월 수식이 변함')
                # 신규 월 헤더가 기존 월 다음 컬럼에 추가됨
                self.assertEqual(str(s1.cell(hr, fmc + mcount).value).strip(), newm)
                # 신규 월 셀은 SUMIFS 수식(값 덮어쓰기 아님)
                self.assertTrue(str(s1.cell(fr, fmc + mcount).value).startswith('=SUMIFS('))

            # (4) 신규 부서 없음 → dept_cross / TOP20 / 대시보드 완전 무변경
            for sheet, lay in self.ps.PIVOT_LAYOUT.items():
                if lay['axis'] != 'dept_cross':
                    continue
                s0, s1 = src_wb[sheet], wf[sheet]
                diffs = sum(1 for rr in range(1, s0.max_row + 1)
                            for cc in range(1, s0.max_column + 1)
                            if s0.cell(rr, cc).value != s1.cell(rr, cc).value)
                self.assertEqual(diffs, 0, f'{sheet} 신규부서 없는데 변경됨')
            # 큰지출_TOP20은 여전히 정적(_apply_pivots/_apply_dashboard 대상 아님)
            if '큰지출_TOP20' in wf.sheetnames:
                s0, s1 = src_wb['큰지출_TOP20'], wf['큰지출_TOP20']
                diffs = sum(1 for rr in range(1, s0.max_row + 1)
                            for cc in range(1, s0.max_column + 1)
                            if s0.cell(rr, cc).value != s1.cell(rr, cc).value)
                self.assertEqual(diffs, 0, '큰지출_TOP20 정적 시트가 변경됨')

            # 요약_대시보드는 _apply_dashboard로 갱신됨: 제목 월범위·월평균분모·추이 신규월
            if '요약_대시보드' in wf.sheetnames:
                dsh = wf['요약_대시보드']
                self.assertIn(f'1~{maxm + 1}월', str(dsh['A1'].value))       # 제목 확장
                self.assertTrue(str(dsh['G5'].value).rstrip().endswith(f'/{len(piv["월합"])}'))  # 분모=월수
                tmonths = []
                r = 18
                while True:
                    v = dsh.cell(r, 1).value
                    if v is None or (isinstance(v, str) and v.startswith('시트 가이드')):
                        break
                    tmonths.append(str(v).strip()); r += 1
                self.assertIn(newm, tmonths)                                # 신규 월 추이행
                self.assertEqual(len(tmonths), len(piv['월합']))            # 추이=존재 월 전체

            # (5) 멱등: 같은 pivots 재적용해도 신규 월 중복 추가 없음
            before = [wf['부서별_월별'].cell(3, c).value for c in range(1, 12)]
            self.ps._apply_pivots(wf, piv)
            after = [wf['부서별_월별'].cell(3, c).value for c in range(1, 12)]
            self.assertEqual(before, after)


if __name__ == '__main__':
    unittest.main()
