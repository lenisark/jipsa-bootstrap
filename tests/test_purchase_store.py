import unittest, importlib.util, tempfile
from pathlib import Path
import openpyxl
SRC = Path(__file__).resolve().parents[1] / 'templates/scripts/slack-jipsa/purchase_store.py'
def load():
    spec = importlib.util.spec_from_file_location('pstore_uut', SRC)
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m); return m

class LogIOTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_create_and_append(self):
        with tempfile.TemporaryDirectory() as d:
            out = Path(d)/'비품 구매기록_자동.xlsx'
            recs=[{'월':'8월','일자':'2026-08-06','부서':'재무회계','용도':'사내비품','카테고리':'IT·전자',
                   '품목':'무선마우스','수량':2,'단가':49900,'금액':99800,'블록ID':'202608#1'}]
            n = self.m.append_purchase_log(out, recs)
            self.assertEqual(n, 1)
            wb = openpyxl.load_workbook(out); ws = wb[self.m.LOG_SHEET]
            self.assertEqual([c.value for c in ws[1]], self.m.LOG_HEADERS)   # 헤더 r1
            self.assertEqual([c.value for c in ws[2]][:10],
                ['8월','2026-08-06','재무회계','사내비품','IT·전자','무선마우스',2,49900,99800,'202608#1'])
            wb.close()
    def test_append_accumulates(self):
        with tempfile.TemporaryDirectory() as d:
            out = Path(d)/'log.xlsx'
            self.m.append_purchase_log(out, [{'월':'8월','품목':'a','금액':1,'블록ID':'202608#1'}])
            self.m.append_purchase_log(out, [{'월':'8월','품목':'b','금액':2,'블록ID':'202608#2'}])
            rows = self.m.read_purchase_log(out)
            self.assertEqual([r['품목'] for r in rows], ['a','b'])   # 누적
    def test_read_roundtrip(self):
        with tempfile.TemporaryDirectory() as d:
            out = Path(d)/'log.xlsx'
            self.m.append_purchase_log(out, [
                {'월':'8월','일자':'2026-08-06','부서':'인사총무','용도':'사내비품','카테고리':'사무용품',
                 '품목':'볼펜','수량':1,'단가':1000,'금액':1000,'블록ID':'202608#1'}])
            rows = self.m.read_purchase_log(out)
            self.assertEqual(len(rows), 1)
            self.assertEqual(set(rows[0].keys()), set(self.m.LOG_HEADERS))
            self.assertEqual(rows[0]['금액'], 1000)
    def test_read_missing_file(self):
        with tempfile.TemporaryDirectory() as d:
            self.assertEqual(self.m.read_purchase_log(Path(d)/'nope.xlsx'), [])

class IntegratedTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_read_integrated(self):
        import tempfile, openpyxl
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            p=_P(d)/'a.xlsx'; wb=openpyxl.Workbook(); ws=wb.active; ws.title='통합원본'
            ws.append(['월','일자','부서','용도','카테고리','품목','금액','블록ID'])
            ws.append(['1월','6일','인사총무','사내비품','사무용품','볼펜',1000,'1월#1'])
            wb.save(p); wb.close()
            rows=self.m.read_integrated(p)
            self.assertEqual(rows[0]['부서'],'인사총무'); self.assertEqual(rows[0]['금액'],1000)

class DashboardTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()

    @staticmethod
    def _make(path):
        wb = openpyxl.Workbook(); w0 = wb.active; w0.title = '통합원본'
        w0.append(['월','일자','부서','용도','카테고리','품목','금액','블록ID'])
        ws = wb.create_sheet('요약_대시보드')
        ws['A1'] = 'DSAuto 비품 지출 요약 — 2026년 1~5월 (v1.2: 5월 추가)'
        ws['A4'] = '1~4월 총 지출'
        ws['G5'] = '=SUMIF(통합원본!D:D,"사내비품",통합원본!G:G)/5'
        for i, dept in enumerate(['전부서 공통','인사총무','관리사무소','영업본부','행정지원파트']):
            r = 9 + i
            ws.cell(r,1).value = i+1; ws.cell(r,2).value = dept
            ws.cell(r,3).value = f'=SUMIFS(통합원본!G:G,통합원본!C:C,"{dept}",통합원본!D:D,"사내비품")'
            ws.cell(r,4).value = f'=C{r}/SUMIF(통합원본!D:D,"사내비품",통합원본!G:G)'
        ws['A17']='월'; ws['B17']='사내비품'; ws['C17']='전월대비'; ws['D17']='참고: 재판매·대고객'
        for i, mo in enumerate(['1월','2월','3월','4월','5월']):
            r = 18 + i; ws.cell(r,1).value = mo
            ws.cell(r,2).value = f'=SUMIFS(통합원본!G:G,통합원본!A:A,"{mo}",통합원본!D:D,"사내비품")'
            ws.cell(r,3).value = '-' if i==0 else f'=B{r}/B{r-1}-1'
        ws['A23'] = '시트 가이드: ①통합원본 ②부서별_월별 …'
        ws.merge_cells('A1:H1'); ws.merge_cells('A23:H23')   # 제목·가이드 행 병합(실파일 모사)
        wb.save(path); wb.close()

    def test_apply_dashboard(self):
        with tempfile.TemporaryDirectory() as d:
            p = Path(d)/'a.xlsx'; self._make(p)
            # 1~8월 존재, 사내비품 부서합: 관리사무소가 최상위가 되도록 구성
            pivots = {'월합': {f'{n}월': 1000 for n in range(1,9)},
                      '부서용도': {('관리사무소','사내비품'): 9000, ('전부서 공통','사내비품'): 5000,
                                   ('인사총무','사내비품'): 3000, ('CX파트','사내비품'): 2000,
                                   ('영업본부','사내비품'): 1000, ('매입부','사내비품'): 500,
                                   ('관리사무소','재판매·대고객'): 100}}
            wb = openpyxl.load_workbook(p); self.m._apply_dashboard(wb, pivots)
            wb.save(p); wb.close()
            wb = openpyxl.load_workbook(p); ws = wb['요약_대시보드']
            self.assertIn('1~8월', ws['A1'].value)                 # 제목 월범위
            self.assertIn('1~8월', ws['A4'].value)                 # 라벨 월범위
            self.assertTrue(ws['G5'].value.rstrip().endswith('/8'))# 월평균 분모 = 월 수
            self.assertEqual(ws['B9'].value, '관리사무소')          # TOP5 재정렬(최상위)
            self.assertEqual(ws['B10'].value, '전부서 공통')
            self.assertIn('관리사무소', ws['C9'].value)             # SUMIFS도 재작성
            self.assertEqual([ws.cell(18+i,1).value for i in range(8)],
                             ['1월','2월','3월','4월','5월','6월','7월','8월'])   # 추이 8개월
            self.assertEqual(ws['C18'].value, '-')                 # 첫 달 전월대비 '-'
            self.assertTrue(str(ws['B18'].value).startswith('=SUMIFS'))
            self.assertEqual(ws.cell(26,1).value[:6], '시트 가이드')  # 가이드 행 재배치(8개월 → r26)
            self.assertIn('A26:H26', [str(x) for x in ws.merged_cells.ranges])  # 가이드 병합 재적용
            self.assertNotIn('A23:H23', [str(x) for x in ws.merged_cells.ranges])  # 옛 위치 병합 해제
            wb.close()


class Top20Test(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_apply_top20(self):
        with tempfile.TemporaryDirectory() as d:
            p = Path(d)/'a.xlsx'
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = '큰지출_TOP20'
            ws['A1'] = '단건 지출 TOP 20 (이상치 점검용)'
            for c, h in enumerate(['순위','월','부서','용도','카테고리','품목','금액'], start=1):
                ws.cell(3, c).value = h
            ws['A4'] = 99; ws['F4'] = '옛데이터'; ws['G4'] = 1   # 덮어써질 값
            wb.save(p); wb.close()
            pivots = {'top20': [(500,'3월','매입부','재판매·대고객','청소·위생','방향제'),
                                (300,'1월','인사총무','사내비품','사무용품','A4'),
                                (100,'2월','전부서 공통','사내비품','다과·음료','커피')]}
            wb = openpyxl.load_workbook(p); self.m._apply_top20(wb, pivots); wb.save(p); wb.close()
            ws = openpyxl.load_workbook(p)['큰지출_TOP20']
            self.assertEqual([ws.cell(4,c).value for c in range(1,8)],
                             [1,'3월','매입부','재판매·대고객','청소·위생','방향제',500])
            self.assertEqual(ws.cell(5,3).value, '인사총무')   # 2위 부서
            self.assertEqual(ws.cell(6,7).value, 100)          # 3위 금액
            self.assertIsNone(ws.cell(7,1).value)              # 3건뿐 → 이후 비움


class MergeTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_latest_analysis(self):
        import tempfile, openpyxl
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            for name in ['260101-HGA-비품주문분석-v1.2.xlsx','260501-HGA-비품주문분석-v1.10.xlsx']:
                wb=openpyxl.Workbook(); wb.save(d/name); wb.close()
            p,(maj,mn)=self.m.latest_analysis(d,'비품주문분석')
            self.assertEqual((maj,mn),(1,10))            # 1.10 > 1.2
    def test_append_integrated(self):
        import tempfile, openpyxl
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d); src=d/'src.xlsx'; out=d/'out.xlsx'
            wb=openpyxl.Workbook(); ws=wb.active; ws.title='통합원본'
            ws.append(['월','일자','부서','용도','카테고리','품목','금액','블록ID'])
            ws.append(['1월','6일','인사총무','사내비품','사무용품','볼펜',1000,'1월#1'])
            wb.save(src); wb.close()
            new=[{'월':'8월','일자':'6일','부서':'재무회계','용도':'사내비품','카테고리':'IT·전자','품목':'마우스','금액':99800,'블록ID':'202608#1'}]
            self.m.write_merged_analysis(src, out, new, pivots=None)
            wb=openpyxl.load_workbook(out); ws=wb['통합원본']
            self.assertEqual(ws.max_row, 3); self.assertEqual(ws['A3'].value,'8월')
