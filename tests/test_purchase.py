import unittest, importlib.util
from pathlib import Path
SRC = Path(__file__).resolve().parents[1] / 'templates/scripts/slack-jipsa/purchase.py'
def load():
    spec = importlib.util.spec_from_file_location('purchase_uut', SRC)
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m); return m
KNOWN = ['전부서 공통','인사총무','매입부','영업본부','영업시스템본부','CX파트']

class DeptTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_exact(self):
        self.assertEqual(self.m.normalize_dept('매입부', KNOWN, {}), ('매입부', False))
    def test_alias(self):
        self.assertEqual(self.m.normalize_dept('영업1본부', KNOWN, {'영업1본부':'영업본부'}), ('영업본부', False))
    def test_contains(self):
        self.assertEqual(self.m.normalize_dept(' 전부서 ', KNOWN, {}), ('전부서 공통', False))
    def test_new(self):
        self.assertEqual(self.m.normalize_dept('신설팀', KNOWN, {}), ('신설팀', True))

class ClassifyTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_batch_one_call(self):
        calls=[]
        def fake(p):
            calls.append(1)
            return ('네 [{"품목":"물티슈, 80개","용도":"사내비품","카테고리":"청소·위생"},'
                    '{"품목":"무선마우스","용도":"사내비품","카테고리":"IT·전자"}] 끝')
        res = self.m.classify_items(['물티슈, 80개','무선마우스'], fake)
        self.assertEqual(len(calls), 1)
        self.assertEqual(res['무선마우스']['카테고리'], 'IT·전자')
    def test_cache_hit_no_llm(self):
        cache = {'물티슈, 80개': {'용도':'사내비품','카테고리':'청소·위생'}}
        def boom(p): raise AssertionError('LLM 호출되면 안 됨')
        res, newc = self.m.classify_with_cache(['물티슈, 80개'], cache, boom)
        self.assertEqual(res['물티슈, 80개']['카테고리'], '청소·위생')
    def test_empty(self):
        self.assertEqual(self.m.classify_items([], lambda p:'x'), {})

class BuildTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_build_basic(self):
        rows = [{'품명':'무선마우스','수량':2,'금액':99800,'부서':'재무회계'}]
        cmap = {'무선마우스': {'용도':'사내비품','카테고리':'IT·전자'}}
        recs, warns = self.m.build_records(rows, cmap, '2026-08-06', ['재무회계'], {})
        r = recs[0]
        self.assertEqual((r['품목'], r['수량'], r['금액'], r['단가']), ('무선마우스',2,99800,49900))
        self.assertEqual((r['용도'], r['카테고리'], r['부서'], r['일자']),
                         ('사내비품','IT·전자','재무회계','2026-08-06'))
    def test_new_dept_warns(self):
        rows=[{'품명':'볼펜','수량':1,'금액':1000,'부서':'미지의팀'}]
        recs, warns = self.m.build_records(rows, {'볼펜':{'용도':'사내비품','카테고리':'사무용품'}},
                                           '2026-08-06', ['인사총무'], {})
        self.assertTrue(any('미지의팀' in w for w in warns))
    def test_zero_qty_unitprice_zero(self):
        recs,_ = self.m.build_records([{'품명':'X','수량':0,'금액':500,'부서':''}],
                                      {'X':{'용도':'사내비품','카테고리':'기타'}}, '2026-08-06', [], {})
        self.assertEqual(recs[0]['단가'], 0)

class ApplyTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_apply_uses_cache_and_writes_log(self):
        import tempfile, json as _j, openpyxl
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d = _P(d)
            cache = d/'c.json'; cache.write_text(_j.dumps({'볼펜':{'용도':'사내비품','카테고리':'사무용품'}}), encoding='utf-8')
            cfg = {'folder':str(d),'purchase_log':'비품 구매기록_자동.xlsx','classify_cache':str(cache),
                   'known_depts':['전부서 공통'],'dept_aliases':{}}
            rows=[{'품명':'볼펜','수량':3,'금액':3000,'부서':'전부서'}]
            def boom(p): raise AssertionError('캐시 적중인데 LLM 호출')
            res = self.m.apply_purchase_record(cfg, rows, boom, '2026-08-06')
            self.assertEqual(res['appended'], 1)
            self.assertIsNone(res['error'])
            r = res['records'][0]
            self.assertEqual((r['월'], r['품목'], r['단가'], r['블록ID']), ('8월','볼펜',1000,'202608#1'))
            wb = openpyxl.load_workbook(d/'비품 구매기록_자동.xlsx'); ws = wb.active
            self.assertEqual(ws.cell(2, 6).value, '볼펜')       # 품목 = 6열, 데이터 r2
            wb.close()
    def test_blockid_seq_continues_same_month(self):
        import tempfile
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            cfg = {'folder':str(d),'purchase_log':'log.xlsx','classify_cache':str(d/'c.json'),
                   'known_depts':['전부서 공통'],'dept_aliases':{}}
            def fake(p): return '[{"품목":"볼펜","용도":"사내비품","카테고리":"사무용품"}]'
            self.m.apply_purchase_record(cfg, [{'품명':'볼펜','수량':1,'금액':1000,'부서':'전부서'}], fake, '2026-08-06')
            res = self.m.apply_purchase_record(cfg, [{'품명':'볼펜','수량':1,'금액':1000,'부서':'전부서'}], fake, '2026-08-07')
            self.assertEqual(res['records'][0]['블록ID'], '202608#2')   # 같은 달 seq 이어짐

class PivotTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_pivots(self):
        rows=[{'월':'1월','부서':'인사총무','용도':'사내비품','카테고리':'사무용품','품목':'볼펜','금액':1000},
              {'월':'1월','부서':'인사총무','용도':'사내비품','카테고리':'사무용품','품목':'A4','금액':2000},
              {'월':'2월','부서':'매입부','용도':'공용(사내·외부 혼재)','카테고리':'청소·위생','품목':'물티슈','금액':5000}]
        p = self.m.compute_pivots(rows)
        self.assertEqual(p['부서월'][('인사총무','1월')], 3000)
        self.assertEqual(p['총계'], 8000)
        self.assertEqual(p['월합']['2월'], 5000)
        self.assertEqual(p['top20'][0][0], 5000)      # 금액 내림차순

class ReflectTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()
    def test_log_to_integrated(self):
        logs=[{'월':'8월','일자':'2026-08-06','부서':'재무회계','용도':'사내비품','카테고리':'IT·전자',
               '품목':'무선마우스','수량':2,'단가':49900,'금액':99800,'블록ID':'202608#1'}]
        out=self.m.log_rows_to_integrated(logs)
        self.assertEqual(out[0]['월'],'8월'); self.assertEqual(out[0]['금액'],99800)
        self.assertEqual(out[0]['일자'],'6일'); self.assertEqual(out[0]['블록ID'],'202608#1')
        self.assertNotIn('수량', out[0])          # 통합원본엔 수량 없음
    def test_unreflected(self):
        integ=[{'월':'5월'}]
        self.assertTrue(self.m.unreflected_months(integ,'8월'))
        self.assertFalse(self.m.unreflected_months(integ,'5월'))

class MergeOrchTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls): cls.m = load()

    @staticmethod
    def _make_analysis(path, integ_rows):
        import openpyxl
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='통합원본'
        ws.append(['월','일자','부서','용도','카테고리','품목','금액','블록ID'])
        for r in integ_rows:
            ws.append(r)
        wb.save(path); wb.close()

    @staticmethod
    def _make_log(path, log_rows):
        import openpyxl
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='기록'
        ws.append(['월','일자','부서','용도','카테고리','품목','수량','단가','금액','블록ID'])   # r1 헤더
        for r in log_rows:
            ws.append(r)
        wb.save(path); wb.close()

    def test_nothing_when_already_reflected(self):
        import tempfile
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            a=d/'260501-HGA-비품주문분석-v1.2.xlsx'
            self._make_analysis(a, [['5월','6일','인사총무','사내비품','사무용품','볼펜',1000,'5월#1']])
            cfg={'folder':str(d),'analysis_prefix':'비품주문분석','dept_code':'HGA',
                 'purchase_log':'비품 구매기록_자동.xlsx'}
            res=self.m.merge_month_into_analysis(cfg,'202605','2026-08-06',dry_run=True)
            self.assertEqual(res['status'],'nothing')     # 5월 이미 반영

    def test_nothing_when_log_missing(self):
        import tempfile
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            a=d/'260501-HGA-비품주문분석-v1.2.xlsx'
            self._make_analysis(a, [['5월','6일','인사총무','사내비품','사무용품','볼펜',1000,'5월#1']])
            cfg={'folder':str(d),'analysis_prefix':'비품주문분석','dept_code':'HGA',
                 'purchase_log':'비품 구매기록_자동.xlsx'}
            res=self.m.merge_month_into_analysis(cfg,'202606','2026-08-06',dry_run=True)
            self.assertEqual(res['status'],'nothing')     # 로그 파일 없음

    def test_nothing_when_month_absent_in_log(self):
        import tempfile
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            a=d/'260501-HGA-비품주문분석-v1.2.xlsx'
            self._make_analysis(a, [['5월','6일','인사총무','사내비품','사무용품','볼펜',1000,'5월#1']])
            self._make_log(d/'비품 구매기록_자동.xlsx',
                [['7월','2026-07-06','매입부','사내비품','IT·전자','마우스',1,50000,50000,'202607#1']])
            cfg={'folder':str(d),'analysis_prefix':'비품주문분석','dept_code':'HGA',
                 'purchase_log':'비품 구매기록_자동.xlsx'}
            res=self.m.merge_month_into_analysis(cfg,'202606','2026-08-06',dry_run=True)
            self.assertEqual(res['status'],'nothing')     # 로그에 6월 없음

    def test_proposed_no_write(self):
        import tempfile, openpyxl
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            a=d/'260501-HGA-비품주문분석-v1.2.xlsx'
            self._make_analysis(a, [['5월','6일','인사총무','사내비품','사무용품','볼펜',1000,'5월#1']])
            self._make_log(d/'비품 구매기록_자동.xlsx',
                [['6월','2026-06-06','매입부','사내비품','IT·전자','마우스',1,50000,50000,'202606#1']])
            cfg={'folder':str(d),'analysis_prefix':'비품주문분석','dept_code':'HGA',
                 'purchase_log':'비품 구매기록_자동.xlsx'}
            before=sorted(p.name for p in d.iterdir())
            res=self.m.merge_month_into_analysis(cfg,'202606','2026-08-06',dry_run=True)
            self.assertEqual(res['status'],'proposed')
            self.assertEqual(res['rows'],1)
            self.assertIsNone(res['out'])
            self.assertEqual(res['summary']['월합']['6월'],50000)
            self.assertEqual(sorted(p.name for p in d.iterdir()), before)   # 파일 무변화

    def test_merged_writes_new_version_only(self):
        import tempfile, openpyxl
        from pathlib import Path as _P
        with tempfile.TemporaryDirectory() as d:
            d=_P(d)
            a=d/'260501-HGA-비품주문분석-v1.2.xlsx'
            self._make_analysis(a, [['5월','6일','인사총무','사내비품','사무용품','볼펜',1000,'5월#1']])
            src_mtime=a.stat().st_mtime
            self._make_log(d/'비품 구매기록_자동.xlsx',
                [['6월','2026-06-06','매입부','사내비품','IT·전자','마우스',1,50000,50000,'202606#1']])
            cfg={'folder':str(d),'analysis_prefix':'비품주문분석','dept_code':'HGA',
                 'purchase_log':'비품 구매기록_자동.xlsx'}
            res=self.m.merge_month_into_analysis(cfg,'202606','2026-08-06',dry_run=False)
            self.assertEqual(res['status'],'merged')
            out=_P(res['out'])
            self.assertEqual(out.name,'260806-HGA-비품주문분석-v1.3.xlsx')   # next ver + when yymmdd
            self.assertTrue(out.exists())
            self.assertEqual(a.stat().st_mtime, src_mtime)                  # 원본 불변
            wb=openpyxl.load_workbook(out); ws=wb['통합원본']
            self.assertEqual(ws.max_row,3); self.assertEqual(ws['A3'].value,'6월')

    def test_yymmdd(self):
        self.assertEqual(self.m._yymmdd('2026-08-06'),'260806')

if __name__ == '__main__':
    unittest.main()
