"""비품관리 xlsx 저장소 — openpyxl read/write, 원자적 저장, 엑셀 잠금감지.

순수 IO만 담당(슬랙/LLM 모름). daemon 형제 모듈.
재고/별칭은 한 파일(비품재고_현황.xlsx)의 두 시트, 이력은 별도 파일.
"""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl

STOCK_SHEET = '재고'
ALIAS_SHEET = '별칭'
STOCK_HEADERS = ['품목', '카테고리', '현재수량', '최소수량', '단위', '비고']
ALIAS_HEADERS = ['원문품목', 'canonical품목', '출처', '확신도', '결정방식', '결정시각']
LEDGER_HEADERS = ['일시', '유형', 'canonical품목', '원문품목', '수량', '처리후잔여',
                  '신청자/발주처', '출처키']


def is_locked(path: Path) -> bool:
    """엑셀이 파일을 열고 있으면 같은 폴더에 '~$이름' 잠금파일이 생긴다."""
    p = Path(path)
    return (p.parent / ('~$' + p.name)).exists()


def _atomic_save(wb, path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + '.tmp')
    wb.save(tmp)
    os.replace(tmp, path)


def _to_int(v) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def read_stock(path: Path) -> dict:
    """재고 시트 → {품목: row dict}. 파일 없으면 {}."""
    path = Path(path)
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if STOCK_SHEET not in wb.sheetnames:
        wb.close(); return {}
    ws = wb[STOCK_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for r in rows[1:]:                      # 첫 행 헤더
        if not r or not r[0]:
            continue
        d = {h: (r[i] if i < len(r) else None) for i, h in enumerate(STOCK_HEADERS)}
        d['현재수량'] = _to_int(d['현재수량'])
        d['최소수량'] = _to_int(d['최소수량'])
        d['품목'] = str(d['품목'])
        d['카테고리'] = d['카테고리'] or ''
        d['단위'] = d['단위'] or ''
        d['비고'] = d['비고'] or ''
        out[d['품목']] = d
    return out


def write_stock(path: Path, rows: dict) -> None:
    """{품목: row} 를 재고 시트로 원자적 저장. 별칭 시트가 있으면 보존."""
    path = Path(path)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        if STOCK_SHEET in wb.sheetnames:
            del wb[STOCK_SHEET]
        ws = wb.create_sheet(STOCK_SHEET, 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = STOCK_SHEET
    ws.append(STOCK_HEADERS)
    for item in sorted(rows):
        d = rows[item]
        ws.append([d.get(h, '') for h in STOCK_HEADERS])
    _atomic_save(wb, path)
