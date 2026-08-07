"""비품 구매기록·분석 xlsx IO. supply_store의 잠금감지/원자저장 재활용."""
from __future__ import annotations
import importlib.util
from pathlib import Path
import openpyxl

_ss = importlib.util.spec_from_file_location('supply_store_io',
        Path(__file__).with_name('supply_store.py'))
supply_store = importlib.util.module_from_spec(_ss); _ss.loader.exec_module(supply_store)
is_locked = supply_store.is_locked
_atomic_save = supply_store._atomic_save

# 데몬 전용 누적 구매로그(단순 플랫 — 수식/합계/요약 없음, 사람 양식과 분리).
LOG_HEADERS = ['월','일자','부서','용도','카테고리','품목','수량','단가','금액','블록ID']
LOG_SHEET = '기록'

def append_purchase_log(log_path, records) -> int:
    """누적 구매로그에 append. 없으면 헤더 생성. 엑셀 열림이면 RuntimeError('locked')."""
    log_path = Path(log_path)
    if is_locked(log_path):
        raise RuntimeError('locked')
    if log_path.exists():
        wb = openpyxl.load_workbook(log_path)
        ws = wb[LOG_SHEET] if LOG_SHEET in wb.sheetnames else wb.create_sheet(LOG_SHEET)
        if ws.cell(row=1, column=1).value in (None, ''):   # 헤더 없으면 추가
            ws.append(LOG_HEADERS)
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = LOG_SHEET
        ws.append(LOG_HEADERS)
    n = 0
    for rec in records:
        ws.append([rec.get(h, '') for h in LOG_HEADERS])
        n += 1
    _atomic_save(wb, log_path)
    return n

def read_purchase_log(log_path) -> list:
    """누적 구매로그(헤더 r1, 데이터 r2~) → LOG_HEADERS dict 리스트. 없으면 []."""
    log_path = Path(log_path)
    if not log_path.exists():
        return []
    wb = openpyxl.load_workbook(log_path, read_only=True, data_only=True)
    if LOG_SHEET not in wb.sheetnames:
        wb.close(); return []
    raw = list(wb[LOG_SHEET].iter_rows(min_row=2, values_only=True)); wb.close()
    out = []
    for r in raw:
        if not r or all(c in (None, '') for c in r):
            continue
        rec = {h: (r[i] if i < len(r) else None) for i, h in enumerate(LOG_HEADERS)}
        if rec.get('품목') in (None, ''):
            continue
        out.append(rec)
    return out

INTEGRATED_HEADERS = ['월','일자','부서','용도','카테고리','품목','금액','블록ID']
INTEGRATED_SHEET = '통합원본'

def read_integrated(analysis_path) -> list:
    wb = openpyxl.load_workbook(Path(analysis_path), read_only=True, data_only=True)
    ws = wb[INTEGRATED_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True)); wb.close()
    out=[]
    for r in rows:
        if not r or r[0] in (None,''): continue
        out.append({h:(r[i] if i < len(r) else None) for i,h in enumerate(INTEGRATED_HEADERS)})
    return out

import re as _re
def latest_analysis(folder, prefix):
    folder = Path(folder); best=None; bestv=(0,0)
    for p in folder.glob(f'*{prefix}-v*.xlsx'):
        m=_re.search(r'-v(\d+)\.(\d+)\.xlsx$', p.name)
        if not m: continue
        v=(int(m.group(1)),int(m.group(2)))
        if v>bestv: bestv, best = v, p
    return best, bestv

def next_version(v): return (v[0], v[1]+1)

def write_merged_analysis(src_path, out_path, new_integrated_rows, pivots=None):
    wb = openpyxl.load_workbook(Path(src_path))
    ws = wb[INTEGRATED_SHEET]
    # 통합원본은 후행 빈 행이 없어(실측: max_row == A열 마지막 데이터행) max_row+1 append로 충분.
    last = ws.max_row
    for i,r in enumerate(new_integrated_rows, start=last+1):
        for ci,h in enumerate(INTEGRATED_HEADERS, start=1):
            ws.cell(row=i, column=ci).value = r.get(h,'')
    if pivots is not None:
        _apply_pivots(wb, pivots)     # PIVOT_LAYOUT 기반 파생시트 확장(Task 12)
        _apply_dashboard(wb, pivots)  # 요약_대시보드 갱신(제목·월평균분모·TOP5·월별추이)
        _apply_top20(wb, pivots)      # 큰지출_TOP20 재생성(전 기간 단건 상위 20)
    _atomic_save(wb, Path(out_path))

# ── 파생 시트 좌표 상수화(실파일 v1.2 실측) + SUMIFS 보존형 확장 ─────────────
# 모든 파생 피벗은 통합원본을 참조하는 전열(全列) SUMIFS 수식이다. 통합원본에
# 행을 append 하면 엑셀이 열 때 자동 재계산되므로, 기존 월/부서 셀은 절대 값으로
# 덮어쓰지 않는다(수식 파괴 금지). _apply_pivots 는 (a) 신규 '월' 열 추가,
# (b) 신규 '부서' 행 추가만 수행하며, 기존 1~5월 열/수식은 손대지 않는다
# → "5월까지 합계 불변"이 구조적으로 보장된다.
#
# axis='month'      : 행=라벨(부서/카테고리/용도), 열=월. 신규 월 컬럼을 SUMIFS로 추가.
# axis='dept_cross' : 행=부서, 열=고정(용도/카테고리). 월 필터 없음 → 기존 부서는
#                     전열 SUMIFS 자동재계산, 신규 부서 행만 추가.
PIVOT_LAYOUT = {
    '부서별_월별':    {'axis':'month', 'header_row':3, 'first_row':4, 'first_month_col':2,
                       'row_axis':'부서', 'row_filter':'C', 'pivot':'부서월'},
    '카테고리별_월별': {'axis':'month', 'header_row':3, 'first_row':4, 'first_month_col':2,
                       'row_axis':'카테고리', 'row_filter':'E', 'pivot':'카테고리월'},
    '용도별_월별':    {'axis':'month', 'header_row':3, 'first_row':4, 'first_month_col':2,
                       'row_axis':'용도', 'row_filter':'D', 'pivot':'용도월'},
    '부서x용도':      {'axis':'dept_cross', 'header_row':3, 'first_row':4, 'first_col':2,
                       'row_filter':'C', 'col_filter':'D',
                       'col_labels':['사내비품','공용(사내·외부 혼재)','재판매·대고객'],
                       'pivot':'부서용도'},
    '부서x카테고리':  {'axis':'dept_cross', 'header_row':3, 'first_row':4, 'first_col':2,
                       'row_filter':'C', 'col_filter':'E',
                       'col_labels':['사무용품','다과·음료','청소·위생','IT·전자','비품·가구','기타'],
                       'pivot':'부서카테고리'},
}
INTEGRATED_AMT = 'G'   # 통합원본 금액 열
_MONTH_RE = _re.compile(r'^\s*(\d+)\s*월\s*$')

def _col(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)

def _month_num(label):
    m = _MONTH_RE.match(str(label or ''))
    return int(m.group(1)) if m else None

def _copy_cell(src, dst):
    from copy import copy as _cp
    dst.value = src.value
    if getattr(src, 'has_style', False):
        dst.font = _cp(src.font); dst.fill = _cp(src.fill)
        dst.border = _cp(src.border); dst.alignment = _cp(src.alignment)
        dst.number_format = src.number_format; dst.protection = _cp(src.protection)

def _sumifs_month(month, filter_col, label):
    return (f'=SUMIFS(통합원본!{INTEGRATED_AMT}:{INTEGRATED_AMT},'
            f'통합원본!A:A,"{month}",통합원본!{filter_col}:{filter_col},"{label}")')

def _sumifs_cross(row_filter, dept, col_filter, col_label):
    return (f'=SUMIFS(통합원본!{INTEGRATED_AMT}:{INTEGRATED_AMT},'
            f'통합원본!{row_filter}:{row_filter},"{dept}",'
            f'통합원본!{col_filter}:{col_filter},"{col_label}")')

def _locate_total_row(ws, first_row):
    """A열에서 '합계' 행을 찾는다(없으면 마지막 라벨 다음 행)."""
    r = first_row
    while ws.cell(r, 1).value not in (None, ''):
        if str(ws.cell(r, 1).value).strip() == '합계':
            return r
        r += 1
    return r

def _row_labels(ws, first_row, total_row):
    return [str(ws.cell(r, 1).value).strip() for r in range(first_row, total_row)]

def _apply_month_sheet(ws, lay, pivots):
    hr, fr, fmc = lay['header_row'], lay['first_row'], lay['first_month_col']
    filt = lay['row_filter']
    total_row = _locate_total_row(ws, fr)
    labels = _row_labels(ws, fr, total_row)
    new_dept_rows = []
    # 신규 부서 행 먼저 추가(월 열 좌표에 영향 없음)
    if lay['row_axis'] == '부서':
        start, new_depts = _add_new_depts(ws, lay, pivots, total_row, labels)
        new_dept_rows = list(range(start, start + len(new_depts)))
        total_row = _locate_total_row(ws, fr)
        labels = _row_labels(ws, fr, total_row)
    # 현재 월 컬럼 스캔
    month_cols = []                       # [(col_idx, month_label)]
    total_col = None
    c = fmc
    while True:
        h = ws.cell(hr, c).value
        if h in (None, ''):
            break
        if str(h).strip() == '합계':
            total_col = c; break
        if _month_num(h) is not None:
            month_cols.append((c, str(h).strip()))
        c += 1
    if total_col is None:
        total_col = fmc + len(month_cols)
    ratio_col = total_col + 1             # 비중
    # 신규 부서 행의 기존 월 컬럼(예: 1~5월)도 SUMIFS로 채운다
    style_body0 = ws.cell(fr, fmc)
    for r in new_dept_rows:
        dep = str(ws.cell(r, 1).value).strip()
        for mc, mo in month_cols:
            _copy_cell(style_body0, ws.cell(r, mc))
            ws.cell(r, mc).value = _sumifs_month(mo, filt, dep)
    have = {lbl for _, lbl in month_cols}
    all_months = {mo for mo in pivots.get('월합', {}) if _month_num(mo) is not None}
    new_months = sorted(all_months - have, key=_month_num)
    if not new_months:
        return           # 파이프라인상 '월 병합'은 항상 신규 월을 동반(신규 부서 단독 병합 없음)
    k = len(new_months)
    # 합계/비중 열을 오른쪽으로 k칸 이동(수식은 아래서 전면 재작성하므로 스타일만 보존)
    for src_c in (ratio_col, total_col):
        for r in range(hr, total_row + 1):
            _copy_cell(ws.cell(r, src_c), ws.cell(r, src_c + k))
            ws.cell(r, src_c).value = None
    new_total_col = total_col + k
    new_ratio_col = ratio_col + k
    # 신규 월 컬럼 채우기(헤더 스타일은 기존 월 헤더에서 복제)
    style_hdr = ws.cell(hr, fmc)
    style_body = ws.cell(fr, fmc)
    for j, mo in enumerate(new_months):
        col = total_col + j
        _copy_cell(style_hdr, ws.cell(hr, col)); ws.cell(hr, col).value = mo
        for r in range(fr, total_row):
            _copy_cell(style_body, ws.cell(r, col))
            ws.cell(r, col).value = _sumifs_month(mo, filt, str(ws.cell(r, 1).value).strip())
    first_L, last_L = _col(fmc), _col(new_total_col - 1)
    tot_L = _col(new_total_col)
    # 합계/비중 수식 재작성(기존 월 열은 손대지 않음)
    for r in range(fr, total_row):
        ws.cell(r, new_total_col).value = f'=SUM({first_L}{r}:{last_L}{r})'
        ws.cell(r, new_ratio_col).value = f'={tot_L}{r}/${tot_L}${total_row}'
    # 합계 행
    for c in range(fmc, new_total_col):
        L = _col(c)
        ws.cell(total_row, c).value = f'=SUM({L}{fr}:{L}{total_row-1})'
    ws.cell(total_row, new_total_col).value = f'=SUM({first_L}{total_row}:{last_L}{total_row})'
    ws.cell(total_row, new_ratio_col).value = 1

def _add_new_depts(ws, lay, pivots, total_row, labels):
    """부서 축 시트에 신규 부서 행을 합계행 직전에 삽입.
    반환: (삽입 시작행, 신규부서 리스트). 신규 없으면 (total_row, [])."""
    pv = pivots.get(lay['pivot'], {})
    depts = []
    for key in pv:
        d = key[0] if isinstance(key, tuple) else key
        if d not in depts:
            depts.append(d)
    new_depts = [d for d in depts if d and d not in labels]
    if not new_depts:
        return total_row, []
    ws.insert_rows(total_row, amount=len(new_depts))
    for i, dep in enumerate(new_depts):
        r = total_row + i
        # 라벨 셀 스타일: 기존 첫 데이터행에서 복제
        _copy_cell(ws.cell(lay['first_row'], 1), ws.cell(r, 1)); ws.cell(r, 1).value = dep
    return total_row, new_depts

def _apply_dept_cross(ws, lay, pivots):
    fr, fc = lay['first_row'], lay['first_col']
    total_row = _locate_total_row(ws, fr)
    labels = _row_labels(ws, fr, total_row)
    _add_new_depts(ws, lay, pivots, total_row, labels)
    new_total_row = _locate_total_row(ws, fr)
    if new_total_row == total_row:
        return                                  # 신규 부서 없음 → 전열 SUMIFS 자동재계산
    # 신규 부서 행에 교차 SUMIFS + 합계 채우기
    col_labels = lay['col_labels']; rf = lay['row_filter']; cf = lay['col_filter']
    style_body = ws.cell(fr, fc)
    for r in range(total_row, new_total_row):    # 새로 삽입된 행들
        dep = str(ws.cell(r, 1).value).strip()
        for j, cl in enumerate(col_labels):
            c = fc + j
            _copy_cell(style_body, ws.cell(r, c))
            ws.cell(r, c).value = _sumifs_cross(rf, dep, cf, cl)
        tot_c = fc + len(col_labels)
        L0, L1 = _col(fc), _col(tot_c - 1)
        _copy_cell(ws.cell(fr, tot_c), ws.cell(r, tot_c))
        ws.cell(r, tot_c).value = f'=SUM({L0}{r}:{L1}{r})'
    # 합계 행 SUM 범위 재작성(삽입으로 늘어난 행 포함)
    tot_c = fc + len(col_labels)
    for c in range(fc, tot_c + 1):
        L = _col(c)
        ws.cell(new_total_row, c).value = f'=SUM({L}{fr}:{L}{new_total_row-1})'

def _apply_pivots(wb, pivots):
    """PIVOT_LAYOUT 기반 파생 시트 확장. SUMIFS 수식 보존(값 덮어쓰기 금지).
    - month 축: 신규 월 컬럼 추가(+ 신규 부서 행). 기존 월 열/수식 불변.
    - dept_cross 축: 신규 부서 행만 추가. 기존 부서는 전열 SUMIFS 자동재계산.
    큰지출_TOP20(정적)·요약_대시보드(수동 라벨/전열 집계)는 손대지 않는다."""
    for sheet, lay in PIVOT_LAYOUT.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        if lay['axis'] == 'month':
            _apply_month_sheet(ws, lay, pivots)
        elif lay['axis'] == 'dept_cross':
            _apply_dept_cross(ws, lay, pivots)


# ── 요약_대시보드 갱신 (실파일 v1.2 레이아웃 실측) ──────────────────────────
# 대시보드는 통합원본 참조 SUMIF/SUMIFS(총지출·사내/재판매·카테고리)라 대부분
# 자동 재계산되지만, 아래 4가지는 고정값·구조라 코드로 갱신한다:
#  (1) 제목/라벨의 '1~N월' 범위(A1,A4)  (2) 월평균 분모 G5의 '/N'(월 수)
#  (3) 부서 TOP5(B9~B13) 라벨+SUMIFS 재정렬(전 기간 사내비품 상위 5)
#  (4) 월별 추이(r18~) 월 행을 현재 존재 월 전체로 재작성 + '시트 가이드' 행 재배치
# 값은 전부 SUMIFS 수식으로 써서 엑셀에서 재계산 → 기존 월 수식과 동일 형식.
DASH_SHEET = '요약_대시보드'
DASH_TREND_START = 18          # 월별 추이 데이터 시작 행(헤더 r17)

def _apply_dashboard(wb, pivots):
    if DASH_SHEET not in wb.sheetnames:
        return
    ws = wb[DASH_SHEET]
    months = sorted((m for m in pivots.get('월합', {}) if _month_num(m) is not None),
                    key=_month_num)
    if not months:
        return
    ncount = len(months)
    maxm = _month_num(months[-1])
    # 부서별 사내비품 지출(부서용도 피벗) → TOP5
    dep = {}
    for (d, use), v in pivots.get('부서용도', {}).items():
        if str(use).strip() == '사내비품':
            dep[d] = dep.get(d, 0) + (v or 0)
    top5 = [d for d, _ in sorted(dep.items(), key=lambda x: -x[1])[:5]]
    # (1) 제목/라벨 월범위
    for coord in ('A1', 'A4'):
        v = ws[coord].value
        if isinstance(v, str):
            ws[coord].value = _re.sub(r'\d+\s*~\s*\d+\s*월', f'1~{maxm}월', v)
    # (2) 월평균 분모 G5의 끝 '/N'
    g5 = ws['G5'].value
    if isinstance(g5, str) and g5.startswith('='):
        ws['G5'].value = _re.sub(r'/\s*\d+\s*$', f'/{ncount}', g5)
    # (3) 부서 TOP5 (r9~r13)
    for i in range(5):
        r = 9 + i
        name = top5[i] if i < len(top5) else None
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = name or None
        if name:
            ws.cell(r, 3).value = _sumifs_cross('C', name, 'D', '사내비품')
            ws.cell(r, 4).value = f'=C{r}/SUMIF(통합원본!D:D,"사내비품",통합원본!G:G)'
        else:
            ws.cell(r, 3).value = None
            ws.cell(r, 4).value = None
    # (4) 월별 추이: 존재 월 전체로 재작성 + 가이드(주석) 행 재배치
    # 가이드 행은 '시트 가이드' 텍스트에 의존하지 않고 구조로 탐지한다:
    # r18부터 '월' 라벨 행이 이어지고, 그 다음 비월(非月) 행이 가이드/주석 행.
    # (텍스트 매칭 폴백으로 max_row까지 지우면 문구가 조금만 바뀌어도 조용히 유실됨)
    WRITE_COLS = (1, 2, 3, 4, 7)               # A일자·B사내·C전월대비·D재판매·G공용
    r0 = DASH_TREND_START
    while _month_num(ws.cell(r0, 1).value) is not None:
        r0 += 1
    guide_row = r0 if ws.cell(r0, 1).value not in (None, '') else None
    guide_text = ws.cell(guide_row, 1).value if guide_row else None
    clear_to = guide_row if guide_row else (r0 - 1)
    # 열별 서식 템플릿(클리어 전 캡처): A/B/D/G=r18, C(전월대비 %)=둘째 월 행
    fmt = {1: ws.cell(DASH_TREND_START, 1).number_format,
           2: ws.cell(DASH_TREND_START, 2).number_format,
           3: ws.cell(min(DASH_TREND_START + 1, clear_to), 3).number_format,
           4: ws.cell(DASH_TREND_START, 4).number_format,
           7: ws.cell(DASH_TREND_START, 7).number_format}
    # 추이 영역에 '완전히 포함된' 병합만 해제(위쪽 헤더 병합은 보존). 가이드 폭 보존.
    guide_span = None
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row < DASH_TREND_START or rng.max_row > clear_to:
            continue
        if guide_row and rng.min_row <= guide_row <= rng.max_row:
            guide_span = (rng.min_col, rng.max_col)
        ws.unmerge_cells(str(rng))
    # 재작성 대상 열만 클리어(E/F/H 등 미사용 열 보존)
    for rr in range(DASH_TREND_START, clear_to + 1):
        for c in WRITE_COLS:
            ws.cell(rr, c).value = None
    r = DASH_TREND_START
    for idx, m in enumerate(months):
        ws.cell(r, 1).value = m
        ws.cell(r, 2).value = _sumifs_month(m, 'D', '사내비품')
        ws.cell(r, 3).value = '-' if idx == 0 else f'=B{r}/B{r - 1}-1'
        ws.cell(r, 4).value = _sumifs_month(m, 'D', '재판매·대고객')
        ws.cell(r, 7).value = _sumifs_month(m, 'D', '공용(사내·외부 혼재)')
        for c in (1, 2, 3, 4, 7):
            ws.cell(r, c).number_format = fmt[c]
        r += 1
    if guide_text:
        ws.cell(r, 1).value = guide_text
        if guide_span:                       # 원래 병합 폭으로 새 위치에 재병합
            ws.merge_cells(start_row=r, start_column=guide_span[0],
                           end_row=r, end_column=guide_span[1])


# ── 큰지출_TOP20 재생성 (정적 값 시트) ──────────────────────────────────────
# 헤더 r3: 순위|월|부서|용도|카테고리|품목|금액. 데이터 r4~. compute_pivots['top20']는
# (금액,월,부서,용도,카테고리,품목) 튜플을 금액 내림차순으로 담고 있다. 정적 값 시트라
# 매 병합 시 전 기간 상위 20건으로 다시 써준다(부족하면 남는 행은 비운다).
TOP20_SHEET = '큰지출_TOP20'
TOP20_FIRST_ROW = 4
TOP20_MAX = 20

def _apply_top20(wb, pivots):
    if TOP20_SHEET not in wb.sheetnames:
        return
    ws = wb[TOP20_SHEET]
    top = pivots.get('top20', [])[:TOP20_MAX]
    r = TOP20_FIRST_ROW
    for i, t in enumerate(top):
        amt, mon, dep, use, cat, item = t
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = mon
        ws.cell(r, 3).value = dep
        ws.cell(r, 4).value = use
        ws.cell(r, 5).value = cat
        ws.cell(r, 6).value = item
        ws.cell(r, 7).value = amt
        r += 1
    for rr in range(r, TOP20_FIRST_ROW + TOP20_MAX):   # 20건 미만이면 남는 행 비우기
        for c in range(1, 8):
            ws.cell(rr, c).value = None
